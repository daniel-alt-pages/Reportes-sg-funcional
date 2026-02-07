"""
PROCESAR - Sistema Unificado de Procesamiento
==============================================
Detecta automaticamente CSV o XLSX y genera todos los reportes.

Uso:
    python procesar.py                    # Procesa todo
    python procesar.py --verificar        # Solo verifica
"""
import pandas as pd
import numpy as np
import json
import os
import re
import sys
import math
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from statistics import mean, stdev, median, variance
from collections import defaultdict

# ============================================================================
# CONFIGURACION
# ============================================================================
BASE_DIR = Path(__file__).parent.parent.parent  # Subir 2 niveles: scripts/core -> scripts -> raíz
ENTRADA_DIR = BASE_DIR / "data" / "input"
SALIDA_DIR = BASE_DIR / "output"

# Crear directorios si no existen
ENTRADA_DIR.mkdir(parents=True, exist_ok=True)
SALIDA_DIR.mkdir(parents=True, exist_ok=True)
(ENTRADA_DIR / "respuestas").mkdir(exist_ok=True)
(ENTRADA_DIR / "claves").mkdir(exist_ok=True)

# Colores para Excel
COLORS = {
    'header': '1F4E79',
    'header2': '2E75B6',
    'success': '70AD47',
    'warning': 'FFC000',
    'danger': 'C00000',
    'light': 'F2F2F2',
    'superior': '00B050',
    'alto': '92D050',
    'medio': 'FFEB9C',
    'bajo': 'FFC7CE',
}

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

# ============================================================================
# FUNCIONES DE LECTURA (CSV y XLSX)
# ============================================================================
def leer_archivo(ruta):
    """Lee un archivo CSV o XLSX automaticamente."""
    ruta = Path(ruta)
    if ruta.suffix.lower() == '.xlsx':
        return pd.read_excel(ruta, engine='openpyxl')
    else:
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                return pd.read_csv(ruta, encoding=encoding)
            except:
                continue
        raise Exception(f"No se pudo leer: {ruta}")

def detectar_archivos():
    """Detecta automaticamente los archivos de respuestas y claves."""
    archivos = {
        'respuestas_s1': None,
        'respuestas_s2': None,
        'claves_s1': None,
        'claves_s2': None
    }
    
    print("\n[1/4] Detectando archivos...")
    
    # Buscar en entrada/respuestas y EXCEL/RESPUESTAS
    dirs_respuestas = [
        ENTRADA_DIR / "respuestas",
        BASE_DIR / "EXCEL" / "RESPUESTAS"
    ]
    
    for dir_resp in dirs_respuestas:
        if not dir_resp.exists():
            continue
        for f in dir_resp.iterdir():
            if f.is_file() and f.suffix.lower() in ['.csv', '.xlsx']:
                f_lower = f.name.lower()
                if 'sesion 1' in f_lower or 'sesion1' in f_lower:
                    archivos['respuestas_s1'] = f
                    print(f"      Respuestas S1: {f.name}")
                elif 'sesion 2' in f_lower or 'sesion2' in f_lower:
                    archivos['respuestas_s2'] = f
                    print(f"      Respuestas S2: {f.name}")
    
    # Buscar claves
    dirs_claves = [
        ENTRADA_DIR / "claves",
        BASE_DIR / "EXCEL" / "CLAVES"
    ]
    
    for dir_claves in dirs_claves:
        if not dir_claves.exists():
            continue
        for f in dir_claves.iterdir():
            if f.is_file() and f.suffix.lower() in ['.csv', '.xlsx']:
                try:
                    df = leer_archivo(f)
                    cols = ' '.join(df.columns).lower()
                    if ' s1 ' in cols or 's1 [' in cols:
                        archivos['claves_s1'] = f
                        print(f"      Claves S1: {f.name}")
                    elif ' s2 ' in cols or 's2 [' in cols:
                        archivos['claves_s2'] = f
                        print(f"      Claves S2: {f.name}")
                except:
                    continue
    
    return archivos

# ============================================================================
# FUNCIONES DE PROCESAMIENTO
# ============================================================================
import unicodedata

def normalizar_texto(texto):
    """Quita acentos y convierte a minusculas."""
    texto = unicodedata.normalize('NFD', str(texto))
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    return texto.lower().strip()

def normalizar_nombre(nombre, apellido):
    """Normaliza nombre y apellido para matching."""
    nombre = normalizar_texto(nombre).replace('\n', ' ').replace('  ', ' ')
    apellido = normalizar_texto(apellido).replace('\n', ' ').replace('  ', ' ')
    return f"{nombre}_{apellido}".strip('_')

def normalizar_correo(correo):
    """Normaliza correo electrónico."""
    if pd.isna(correo) or not correo:
        return None
    return str(correo).strip().lower()

def limpiar_id(valor):
    """Limpia y normaliza el numero de identificacion."""
    if pd.isna(valor):
        return None
    val_str = str(valor).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    val_limpio = ''.join(c for c in val_str if c.isdigit())
    return val_limpio if val_limpio else None

def generar_claves_identificacion(correo, nombre, apellido, num_id):
    """Genera múltiples claves de identificación para un estudiante."""
    claves = []
    
    # Clave por correo (más confiable)
    correo_norm = normalizar_correo(correo)
    if correo_norm and '@' in correo_norm:
        claves.append(('correo', correo_norm))
    
    # Clave por nombre+apellido
    nombre_clave = normalizar_nombre(nombre, apellido)
    if nombre_clave and len(nombre_clave) > 3:
        claves.append(('nombre', nombre_clave))
    
    # Clave por número de identificación
    id_limpio = limpiar_id(num_id)
    if id_limpio and len(id_limpio) >= 5:
        claves.append(('id', id_limpio))
    
    return claves

def cargar_calibracion():
    """Carga los datos de calibración (dificultad de preguntas)."""
    ruta_calib = ENTRADA_DIR / "calibracion.json"
    if not ruta_calib.exists():
        return {}
    try:
        with open(ruta_calib, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def cargar_invalidaciones():
    """Carga la configuración de preguntas invalidadas."""
    ruta_config = ENTRADA_DIR / "invalidaciones.json"
    if not ruta_config.exists():
        return []
    
    try:
        with open(ruta_config, 'r', encoding='utf-8') as f:
            config = json.load(f)
        return config.get('invalidaciones', [])
    except Exception as e:
        print(f"      [!] Error al cargar invalidaciones: {e}")
        return []

def es_pregunta_invalidada(numero_pregunta, materia, sesion, invalidaciones):
    """Verifica si una pregunta está invalidada."""
    for inv in invalidaciones:
        if inv.get('sesion', '').upper() == sesion.upper():
            if inv.get('numero_pregunta') == numero_pregunta:
                # Verificar materia si está especificada
                if 'materia' in inv:
                    if inv['materia'].lower() in materia.lower():
                        return True
                else:
                    return True
    return False

def cargar_claves(ruta):
    """Carga las claves de respuestas correctas."""
    df = leer_archivo(ruta)
    claves = {}
    for col in df.columns:
        if '[' in col and ']' in col:
            respuesta = df[col].iloc[0]
            claves[col] = str(respuesta).strip().upper()
    return claves

def extraer_numero_pregunta(nombre_columna):
    """Extrae el numero de pregunta del nombre de columna."""
    match = re.search(r'\[(\d+)\.?\]', nombre_columna)
    return int(match.group(1)) if match else None

def identificar_materia(nombre_columna):
    """Identifica la materia basandose en el nombre de la columna."""
    col_lower = nombre_columna.lower()
    if 'matem' in col_lower:
        return 'matemáticas'
    elif 'lectura' in col_lower or 'critica' in col_lower:
        return 'lectura crítica'
    elif 'sociales' in col_lower or 'ciudadanas' in col_lower:
        return 'sociales y ciudadanas'
    elif 'ciencias' in col_lower or 'naturales' in col_lower:
        return 'ciencias naturales'
    elif 'ingl' in col_lower:
        return 'inglés'
    return None

def procesar_sesion(ruta_respuestas, ruta_claves, nombre_sesion, invalidaciones=None):
    """Procesa una sesion y retorna los resultados por estudiante."""
    if not ruta_respuestas or not ruta_claves:
        return {}
    
    if invalidaciones is None:
        invalidaciones = []
    
    df = leer_archivo(ruta_respuestas)
    claves = cargar_claves(ruta_claves)
    
    # Encontrar columnas
    col_id = col_nombre = col_apellido = col_correo = col_institucion = None
    col_telefono = col_tipo_id = col_tipo_examen = col_departamento = None
    
    for col in df.columns:
        cl = normalizar_texto(col)  # Sin acentos
        if col_correo is None and ('correo' in cl or 'email' in cl):
            col_correo = col
        elif col_id is None and ('numero de identificacion' in cl or 'numero_identificacion' in cl or ('identificacion' in cl and 'tipo' not in cl)):
            col_id = col
        elif col_nombre is None and 'nombre' in cl and ('completo' in cl or 'apellido' not in cl):
            col_nombre = col
        elif col_apellido is None and 'apellido' in cl:
            col_apellido = col
        elif col_institucion is None and 'institucion' in cl:
            col_institucion = col
        elif col_telefono is None and ('telefono' in cl or 'whatsapp' in cl):
            col_telefono = col
        elif col_tipo_id is None and 'tipo' in cl and 'identificacion' in cl:
            col_tipo_id = col
        elif col_tipo_examen is None and 'examen' in cl and 'que' in cl:
            col_tipo_examen = col
        elif col_departamento is None and 'departamento' in cl:
            col_departamento = col
    
    if not col_id:
        print(f"      [!] No se encontro columna de ID en {nombre_sesion}")
        return {}
    
    resultados = {}
    
    for _, row in df.iterrows():
        id_est = limpiar_id(row.get(col_id))
        if not id_est:
            continue
        
        nombre = str(row.get(col_nombre, '')).strip() if col_nombre else ''
        apellido = str(row.get(col_apellido, '')).strip() if col_apellido else ''
        correo = str(row.get(col_correo, '')).strip() if col_correo else ''
        institucion = str(row.get(col_institucion, '')).strip() if col_institucion else 'Sin Asignar'
        telefono = str(row.get(col_telefono, '')).strip() if col_telefono else ''
        tipo_id = str(row.get(col_tipo_id, '')).strip() if col_tipo_id else 'C.C.'
        tipo_examen = str(row.get(col_tipo_examen, '')).strip() if col_tipo_examen else 'ESTUDIANTE'
        departamento = str(row.get(col_departamento, '')).strip() if col_departamento else 'No aplica'
        
        if id_est not in resultados:
            resultados[id_est] = {
                'informacion_personal': {
                    'correo_electronico': correo,
                    'nombres': nombre.replace('\n', ' '),
                    'apellidos': apellido,
                    'nombre_completo': f"{nombre.replace(chr(10), ' ')} {apellido}".strip(),
                    'numero_identificacion': id_est,
                    'institucion': institucion if institucion else 'Sin Asignar',
                    'telefono': telefono if telefono and telefono != 'nan' else '',
                    'tipo_identificacion': tipo_id if tipo_id and tipo_id != 'nan' else 'C.C.',
                    'municipio': departamento if departamento and departamento != 'nan' else 'No aplica'
                },
                'tipo': tipo_examen if tipo_examen and tipo_examen != 'nan' else 'ESTUDIANTE',
                'puntajes': {},
                'respuestas_detalladas': {},
                'sesiones': [],
                's1_aciertos': 0, 's1_total': 0,
                's2_aciertos': 0, 's2_total': 0
            }
        
        if nombre_sesion in resultados[id_est]['sesiones']:
            continue
        
        resultados[id_est]['sesiones'].append(nombre_sesion)
        
        # Procesar respuestas
        for clave_col, resp_correcta in claves.items():
            if clave_col not in df.columns:
                continue
            
            resp_est = str(row.get(clave_col, '')).strip().upper() if pd.notna(row.get(clave_col)) else ''
            materia = identificar_materia(clave_col)
            numero = extraer_numero_pregunta(clave_col)
            
            if materia and numero:
                if materia not in resultados[id_est]['puntajes']:
                    resultados[id_est]['puntajes'][materia] = {'correctas': 0, 'total_preguntas': 0}
                if materia not in resultados[id_est]['respuestas_detalladas']:
                    resultados[id_est]['respuestas_detalladas'][materia] = []
                
                # Verificar si la pregunta está invalidada (NULA - no cuenta para nada)
                pregunta_invalidada = es_pregunta_invalidada(numero, materia, nombre_sesion, invalidaciones)
                
                if pregunta_invalidada:
                    # Pregunta NULA: NO se cuenta en el total, NO suma, NO resta
                    # Solo se registra para referencia pero no afecta puntajes
                    resultados[id_est]['respuestas_detalladas'][materia].append({
                        'numero': numero,
                        'respuesta_estudiante': resp_est,
                        'respuesta_correcta': '*ANULADA*',
                        'es_correcta': None,  # None indica que no aplica
                        'invalidada': True
                    })
                    # NO incrementamos total_preguntas ni correctas
                    continue  # Saltar al siguiente registro
                
                # Pregunta válida: procesar normalmente
                es_correcta = resp_est == resp_correcta
                resultados[id_est]['puntajes'][materia]['total_preguntas'] += 1
                if es_correcta:
                    resultados[id_est]['puntajes'][materia]['correctas'] += 1
                
                resultados[id_est]['respuestas_detalladas'][materia].append({
                    'numero': numero,
                    'respuesta_estudiante': resp_est,
                    'respuesta_correcta': resp_correcta,
                    'es_correcta': es_correcta
                })
        
        # Calcular totales de sesion
        total_correctas = sum(p['correctas'] for p in resultados[id_est]['puntajes'].values())
        total_preguntas = sum(p['total_preguntas'] for p in resultados[id_est]['puntajes'].values())
        
        if nombre_sesion == 'S1':
            resultados[id_est]['s1_aciertos'] = total_correctas
            resultados[id_est]['s1_total'] = total_preguntas
        else:
            resultados[id_est]['s2_aciertos'] = total_correctas
            resultados[id_est]['s2_total'] = total_preguntas
    
    return resultados

def calcular_puntaje_icfes(porcentaje):
    """Calcula el puntaje ICFES (0-100) basado en porcentaje de aciertos."""
    if porcentaje >= 100:
        return 100.0
    elif porcentaje >= 96:
        return 80 + ((porcentaje - 96) / 3.99) * 6
    elif porcentaje >= 75:
        return 65 + ((porcentaje - 75) / 20) * 20
    elif porcentaje >= 50:
        return 45 + ((porcentaje - 50) / 24) * 19
    elif porcentaje >= 25:
        return 30 + ((porcentaje - 25) / 24) * 14
    else:
        return (porcentaje / 25) * 29

def calcular_puntaje_icfes_avanzado(puntaje_data, materia, calibracion):
    """
    Calcula puntaje usando pesos y penalización por inconsistencia.
    """
    # Si no hay calibración para esta materia, usar método antiguo (porcentaje)
    # Verificar si tenemos datos de calibración para preguntas de esta materia
    preguntas_calibradas = [k for k,v in calibracion.items() if v['materia'] == materia]
    
    if not preguntas_calibradas:
        # Fallback: Cálculo lineal simple
        correctas = puntaje_data['correctas']
        total = puntaje_data['total_preguntas']
        porcentaje = (correctas / total * 100) if total > 0 else 0
        return calcular_puntaje_icfes(porcentaje)

    # 1. Calcular Puntaje Base Ponderado
    puntaje_bruto = 0
    max_posible = 0
    
    # Contadores para inconsistencia
    fallo_facil = False
    acierto_dificil = False
    
    # Necesitamos saber qué preguntas ESPECÍFICAS contestó bien/mal
    # Esto requiere que 'puntaje_data' tenga el detalle, pero actualmente se guarda aparte.
    # En el flujo actual, 'calcular_puntajes_finales' itera sobre 'estudiantes'.
    # Los detalles están en 'respuestas_detalladas'.
    # MODIFICAREMOS la llamada para pasar respuestas_detalladas.
    
    return 0 # Placeholder, logic moved to main function to access details

def calcular_puntajes_finales(estudiantes):
    """Calcula puntajes finales y global usando CALIBRACIÓN."""
    calibracion = cargar_calibracion()
    
    PESOS = {
        'matemáticas': 3,
        'lectura crítica': 3,
        'sociales y ciudadanas': 3,
        'ciencias naturales': 3,
        'inglés': 1
    }
    TOTAL_PESOS = sum(PESOS.values())
    
    for id_est, datos in estudiantes.items():
        suma_ponderada = 0
        respuestas_det = datos.get('respuestas_detalladas', {})
        
        for materia, puntaje_data in datos['puntajes'].items():
            # Obtener detalles de respuestas para esta materia
            detalles = respuestas_det.get(materia, [])
            
            # TOTAL REAL DE PREGUNTAS POR MATERIA (del examen ICFES)
            # Las preguntas de sesiones NO presentadas cuentan como INCORRECTAS
            PREGUNTAS_POR_MATERIA = {
                # materia: {sesion: cantidad}
                'lectura crítica': {'S1': 41, 'S2': 0},
                'matemáticas': {'S1': 25, 'S2': 25},
                'sociales y ciudadanas': {'S1': 25, 'S2': 25},
                'ciencias naturales': {'S1': 29, 'S2': 29},
                'inglés': {'S1': 0, 'S2': 55}
            }
            
            sesiones_estudiante = datos.get('sesiones', [])
            tiene_s1 = 'S1' in sesiones_estudiante
            tiene_s2 = 'S2' in sesiones_estudiante
            
            # Obtener estructura de preguntas para esta materia
            estructura = PREGUNTAS_POR_MATERIA.get(materia, {'S1': 0, 'S2': 0})
            preguntas_s1 = estructura['S1']
            preguntas_s2 = estructura['S2']
            total_real = preguntas_s1 + preguntas_s2  # Total de preguntas del examen completo
            
            # Calcular preguntas respondidas y faltantes
            preguntas_respondidas = puntaje_data.get('total_preguntas', 0)
            correctas = puntaje_data.get('correctas', 0)
            
            # Agregar las preguntas de sesiones NO presentadas como INCORRECTAS
            preguntas_faltantes = 0
            if not tiene_s1:
                preguntas_faltantes += preguntas_s1
            if not tiene_s2:
                preguntas_faltantes += preguntas_s2
            
            # El nuevo total incluye las preguntas faltantes (todas incorrectas)
            total_ajustado = preguntas_respondidas + preguntas_faltantes
            # Las correctas siguen siendo las mismas (no hay correctas en sesiones no presentadas)
            
            # Actualizar puntaje_data con los valores ajustados
            puntaje_data['total_preguntas'] = total_ajustado
            puntaje_data['preguntas_faltantes'] = preguntas_faltantes  # Para referencia
            
            puntaje_es_cero = False
            
            # Si no tiene ninguna sesión relevante para esta materia → 0 puntos
            if total_ajustado == 0:
                puntaje_data['puntaje'] = 0
                puntaje_es_cero = True
            # Si tiene 0 correctas → 0 puntos
            elif correctas == 0:
                puntaje_data['puntaje'] = 0
                puntaje_es_cero = True
            
            # Si el puntaje es 0, sumar a ponderación y saltar cálculo normal
            if puntaje_es_cero:
                if materia in PESOS:
                    suma_ponderada += 0  # Explícitamente sumar 0
                continue
            
            # Verificar si existe calibración para esta materia
            tiene_calibracion = any(c['materia'] == materia for c in calibracion.values())
            
            if not tiene_calibracion or not detalles:
                # METODO LEGACY (Sin calibración)
                total = puntaje_data['total_preguntas']
                correctas = puntaje_data['correctas']
                porc = (correctas / total * 100) if total > 0 else 0
                puntaje_final = calcular_puntaje_icfes(porc)
            else:
                # METODO AVANZADO (Con Inconsistencia)
                score_ponderado = 0
                max_score_posible = 0
                
                fallo_faciles = 0
                acierto_dificiles = 0
                total_faciles = 0
                
                for r in detalles:
                    num = r['numero']
                    es_corr = r['es_correcta']
                    if es_corr is None: continue # Anulada
                    
                    # Buscar en calibración
                    key = f"{materia}_{num}"
                    info = calibracion.get(key, {'peso': 1, 'clasificacion': 'MEDIA', 'es_cascara': False})
                    
                    peso = info.get('peso', 1)
                    clasificacion = info.get('clasificacion', info.get('tipo', 'MEDIA'))  # Compatibilidad
                    es_cascara = info.get('es_cascara', 'FACIL' in clasificacion)
                    
                    max_score_posible += peso
                    
                    if es_corr:
                        score_ponderado += peso
                        if 'DIFICIL' in clasificacion:
                            acierto_dificiles += 1
                    else:
                        if es_cascara:
                            fallo_faciles += 1
                            
                    if es_cascara:
                        total_faciles += 1

                # Calcular Porcentaje Ponderado Inicial (0-100)
                if max_score_posible > 0:
                    base_score = (score_ponderado / max_score_posible) * 100
                else:
                    base_score = 0
                    
                # APLICAR PENALIZACIÓN POR INCONSISTENCIA (SUAVIZADA)
                # Regla: Penalización moderada si falla cáscaras y acierta difíciles
                penalizacion = 0
                
                if total_faciles > 0 and fallo_faciles > 0:
                    # Penalización suave por inconsistencia
                    if acierto_dificiles > 0:
                        # -1 punto por cada cáscara fallada si hay inconsistencia
                        penalizacion = fallo_faciles * 1
                        penalizacion = min(penalizacion, 10)  # Máximo 10 puntos
                    else:
                        # Sin inconsistencia: -0.5 puntos por cáscara fallada
                        penalizacion = fallo_faciles * 0.5
                        penalizacion = min(penalizacion, 5)
                         
                puntaje_final = max(0, base_score - penalizacion)

                # TOPES DESLIZANTES (SUAVIZADOS)
                # 0 errores = 100 (máximo)
                # 1 error = tope del área
                # 2+ errores = baja moderadamente
                
                TOPES_1_ERROR = {
                    'matemáticas': 86,
                    'lectura crítica': 82,
                    'sociales y ciudadanas': 84,
                    'ciencias naturales': 82,
                    'inglés': 87
                }
                
                PENALIDAD_POR_ERROR = {
                    'matemáticas': 2,
                    'lectura crítica': 2,
                    'sociales y ciudadanas': 2,
                    'ciencias naturales': 2,
                    'inglés': 2
                }
                
                # Calcular errores con COSTO REDUCIDO según tipo de pregunta
                costo_errores = 0
                for r in detalles:
                    if r.get('es_correcta') == False:
                        key = f"{materia}_{r['numero']}"
                        info = calibracion.get(key, {})
                        clasificacion = info.get('clasificacion', 'MEDIA')
                        
                        # Costos suavizados
                        if 'MUY_FACIL' in clasificacion:
                            costo_errores += 1.3
                        elif 'FACIL' in clasificacion:
                            costo_errores += 1.1
                        elif 'MEDIA' in clasificacion:
                            costo_errores += 1.0
                        elif 'DIFICIL' in clasificacion:
                            costo_errores += 0.7
                        else:  # MUY_DIFICIL
                            costo_errores += 0.5
                
                num_errores_real = len([r for r in detalles if r.get('es_correcta') == False])
                
                tope_1_error = TOPES_1_ERROR.get(materia, 88)
                penal_error = PENALIDAD_POR_ERROR.get(materia, 2)
                
                if num_errores_real == 0:
                    # Examen perfecto: 100 puntos máximo
                    pass
                else:
                    # Tope deslizante suavizado
                    tope_calculado = tope_1_error - ((costo_errores - 1) * penal_error)
                    
                    # CORRECCIÓN SUPERIOR: Si tiene errores, NUNCA puede superar el tope configurado
                    # (Incluso si falló una muy difícil que bonifica)
                    if num_errores_real > 0:
                        tope_calculado = min(tope_calculado, tope_1_error)
                        
                    tope_calculado = max(40, tope_calculado)  # Mínimo 40 puntos
                    puntaje_final = min(puntaje_final, tope_calculado)
            
            puntaje_data['puntaje'] = int(round(puntaje_final))
            
            if materia in PESOS:
                suma_ponderada += puntaje_data['puntaje'] * PESOS[materia]
        
        promedio = suma_ponderada / TOTAL_PESOS
        datos['puntaje_global'] = int(round(promedio * 5))

# ============================================================================
# FUNCIONES DE ANÁLISIS ESTADÍSTICO AVANZADO (Power BI Style)
# ============================================================================

def calcular_racha_consecutiva(respuestas_detalladas):
    """
    Calcula la racha actual de respuestas correctas e incorrectas consecutivas.
    Retorna: (racha_correctas, racha_incorrectas, max_racha_correctas, max_racha_incorrectas)
    """
    if not respuestas_detalladas:
        return 0, 0, 0, 0
    
    # Aplanar todas las respuestas en orden
    todas_respuestas = []
    for materia, respuestas in respuestas_detalladas.items():
        for r in respuestas:
            todas_respuestas.append(r.get('es_correcta', False))
    
    if not todas_respuestas:
        return 0, 0, 0, 0
    
    # Racha actual (desde el final)
    racha_actual = 1
    es_correcta_actual = todas_respuestas[-1]
    for i in range(len(todas_respuestas) - 2, -1, -1):
        if todas_respuestas[i] == es_correcta_actual:
            racha_actual += 1
        else:
            break
    
    racha_correctas_actual = racha_actual if es_correcta_actual else 0
    racha_incorrectas_actual = racha_actual if not es_correcta_actual else 0
    
    # Máximas rachas históricas
    max_correctas = max_incorrectas = 0
    current_correctas = current_incorrectas = 0
    
    for es_correcta in todas_respuestas:
        if es_correcta:
            current_correctas += 1
            current_incorrectas = 0
            max_correctas = max(max_correctas, current_correctas)
        else:
            current_incorrectas += 1
            current_correctas = 0
            max_incorrectas = max(max_incorrectas, current_incorrectas)
    
    return racha_correctas_actual, racha_incorrectas_actual, max_correctas, max_incorrectas

def calcular_estadisticas_estudiante(est_data, todos_puntajes_globales):
    """
    Calcula métricas estadísticas avanzadas para un estudiante.
    Retorna diccionario con: media, desviación, varianza, percentil, IC 95%, rachas
    """
    puntajes = est_data.get('puntajes', {})
    respuestas_det = est_data.get('respuestas_detalladas', {})
    puntaje_global = est_data.get('puntaje_global', 0)
    
    # Extraer puntajes por área
    puntajes_areas = []
    for materia, data in puntajes.items():
        puntaje = data.get('puntaje', 0)
        if puntaje > 0:
            puntajes_areas.append(puntaje)
    
    # Métricas estadísticas basadas en puntajes por área del estudiante
    if len(puntajes_areas) >= 2:
        media_areas = np.mean(puntajes_areas)
        std_areas = np.std(puntajes_areas, ddof=1)  # Sample std
        var_areas = np.var(puntajes_areas, ddof=1)  # Sample variance
        
        # Intervalo de confianza 95% para la media del estudiante
        # Usando valores t aproximados para n pequeño (tabla t de Student)
        n = len(puntajes_areas)
        se = std_areas / np.sqrt(n)
        # Valores t para IC 95% con diferentes grados de libertad
        t_table = {1: 12.71, 2: 4.30, 3: 3.18, 4: 2.78, 5: 2.57, 6: 2.45, 7: 2.36, 8: 2.31, 9: 2.26, 10: 2.23}
        t_val = t_table.get(n-1, 1.96)  # Default a z=1.96 para n grande
        ic_lower = media_areas - t_val * se
        ic_upper = media_areas + t_val * se
    else:
        media_areas = puntajes_areas[0] if puntajes_areas else 0
        std_areas = 0
        var_areas = 0
        ic_lower = media_areas
        ic_upper = media_areas
    
    # Percentil (ranking) basado en puntaje global vs todos los estudiantes
    # Implementación manual de percentileofscore
    if todos_puntajes_globales and puntaje_global > 0:
        menores = sum(1 for p in todos_puntajes_globales if p < puntaje_global)
        iguales = sum(1 for p in todos_puntajes_globales if p == puntaje_global)
        percentil = ((menores + 0.5 * iguales) / len(todos_puntajes_globales)) * 100
    else:
        percentil = 0
    
    # Rachas de respuestas
    racha_corr, racha_inc, max_corr, max_inc = calcular_racha_consecutiva(respuestas_det)
    
    # Totales absolutos
    total_correctas = sum(p.get('correctas', 0) for p in puntajes.values())
    total_preguntas = sum(p.get('total_preguntas', 0) for p in puntajes.values())
    total_incorrectas = total_preguntas - total_correctas
    
    return {
        'media_areas': round(media_areas, 2),
        'std_areas': round(std_areas, 2),
        'var_areas': round(var_areas, 2),
        'percentil': round(percentil, 1),
        'ic_lower': round(ic_lower, 2),
        'ic_upper': round(ic_upper, 2),
        'racha_correctas': racha_corr,
        'racha_incorrectas': racha_inc,
        'max_racha_correctas': max_corr,
        'max_racha_incorrectas': max_inc,
        'total_correctas': total_correctas,
        'total_incorrectas': total_incorrectas,
        'total_preguntas': total_preguntas,
        'tasa_acierto': round((total_correctas / total_preguntas * 100) if total_preguntas > 0 else 0, 1)
    }

def calcular_estadisticas_grupo(estudiantes_list):
    """
    Calcula estadísticas a nivel de grupo para comparación.
    """
    puntajes_globales = [e.get('puntaje_global', 0) for e in estudiantes_list if e.get('puntaje_global', 0) > 0]
    
    if not puntajes_globales:
        return {'media': 0, 'std': 0, 'min': 0, 'max': 0}
    
    return {
        'media': round(np.mean(puntajes_globales), 2),
        'std': round(np.std(puntajes_globales, ddof=1), 2),
        'min': min(puntajes_globales),
        'max': max(puntajes_globales),
        'mediana': round(np.median(puntajes_globales), 2)
    }

def calcular_indice_global_colegio(estudiantes_list):
    """
    Calcula el INDICE GLOBAL del colegio según metodología ICFES (Boletín 12).
    Regla: Promedio ponderado de indices de materias (0-1) usando solo el TOP 80% de estudiantes.
    Fórmula: 3*Mat + 3*Lec + 3*Nat + 3*Soc + 1*Ing  (Todo sobre 13 si se quiere normalizar, pero ICFES suma ponderada)
    El indice por materia es el promedio de los puntajes (0-1) de los estudiantes del top 80%.
    """
    # 1. Filtrar estudiantes con puntaje > 0
    validos = [e for e in estudiantes_list if e.get('puntaje_global', 0) > 0]
    
    if len(validos) < 9:
        return {'indice': 0, 'clasificacion': 'N/A (Min 9 est)', 'top_80_count': len(validos)}
    
    # 2. Ordenar descendente por puntaje global
    validos.sort(key=lambda x: x['puntaje_global'], reverse=True)
    
    # 3. Seleccionar Top 80%
    n_total = len(validos)
    n_top = math.ceil(n_total * 0.8)
    top_80 = validos[:n_top]
    
    if not top_80:
        return {'indice': 0, 'clasificacion': 'N/A', 'top_80_count': 0}

    # 4. Calcular promedios por area para este grupo (convertidos a indice 0-1)
    # Los puntajes en el script están en 0-100.
    def prom_area(key):
        vals = [e.get('puntajes', {}).get(key, {}).get('puntaje', 0) for e in top_80]
        return mean(vals) / 100.0 if vals else 0

    i_mat = prom_area('matemáticas')
    i_lec = prom_area('lectura crítica')
    i_soc = prom_area('sociales y ciudadanas')
    i_nat = prom_area('ciencias naturales')
    i_ing = prom_area('inglés')
    
    # 5. Fórmula Índice Global (Boletín 12)
    # Indice Global = 3*MA + 3*LC + 3*CN + 3*SC + 1*ING
    # El resultado teórico máximo es 3+3+3+3+1 = 13.
    indice_global = (3 * i_mat) + (3 * i_lec) + (3 * i_nat) + (3 * i_soc) + (1 * i_ing)
    
    # Clasificación aproximada (basada en puntos de corte históricos 2021)
    # Estos valores son referenciales, cambian cada año por resolución.
    # Ref: Boletín 12 habla de categorias A+, A, B, C, D.
    cat = 'D'
    if indice_global >= 0.78 * 13: # Aprox A+
        cat = 'A+'
    elif indice_global >= 0.72 * 13:
        cat = 'A'
    elif indice_global >= 0.65 * 13:
        cat = 'B'
    elif indice_global >= 0.58 * 13:
        cat = 'C'
        
    return {
        'indice': round(indice_global, 4),
        'clasificacion': cat,
        'top_80_count': n_top,
        'promedios_top80': {
            'mat': i_mat, 'lec': i_lec, 'nat': i_nat, 'soc': i_soc, 'ing': i_ing 
        }
    }

# ============================================================================
# FUNCIONES DE GENERACION DE REPORTES
# ============================================================================
def calcular_nivel(puntaje, es_global=True):
    """Calcula nivel y color basado en puntaje."""
    if es_global:
        if puntaje >= 400:
            return 'Superior', COLORS['superior']
        elif puntaje >= 325:
            return 'Alto', COLORS['alto']
        elif puntaje >= 250:
            return 'Medio', COLORS['medio']
        else:
            return 'Bajo', COLORS['bajo']
    else:
        if puntaje >= 80:
            return 'Superior', COLORS['superior']
        elif puntaje >= 65:
            return 'Alto', COLORS['alto']
        elif puntaje >= 50:
            return 'Medio', COLORS['medio']
        else:
            return 'Bajo', COLORS['bajo']

def aplicar_header(ws, fila, columnas, color=None):
    """Aplica formato a headers."""
    color = color or COLORS['header']
    for col, texto in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=texto)
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def generar_reporte_mejorado(estudiantes_list):
    """Genera el reporte mejorado con todas las hojas."""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Preparar datos
    datos = []
    for est in estudiantes_list:
        info = est['informacion_personal']
        puntajes = est.get('puntajes', {})
        puntaje_global = est.get('puntaje_global', 0)
        nivel, color = calcular_nivel(puntaje_global)
        
        datos.append({
            'institucion': info.get('institucion', 'Sin Asignar'),
            'correo': info.get('correo_electronico', ''),
            'documento': info.get('numero_identificacion', ''),
            'nombres': info.get('nombres', ''),
            'apellidos': info.get('apellidos', ''),
            'nombre_completo': info.get('nombre_completo', ''),
            's1_aciertos': est.get('s1_aciertos', 0),
            's1_total': est.get('s1_total', 120),
            's2_aciertos': est.get('s2_aciertos', 0),
            's2_total': est.get('s2_total', 134),
            'matematicas': puntajes.get('matemáticas', {}).get('puntaje', 0),
            'lectura': puntajes.get('lectura crítica', {}).get('puntaje', 0),
            'sociales': puntajes.get('sociales y ciudadanas', {}).get('puntaje', 0),
            'ciencias': puntajes.get('ciencias naturales', {}).get('puntaje', 0),
            'ingles': puntajes.get('inglés', {}).get('puntaje', 0),
            'puntaje_global': puntaje_global,
            'nivel': nivel,
            'color': color
        })
    
    # Hoja Resumen
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.merge_cells('A1:D1')
    ws['A1'].value = "REPORTE EJECUTIVO - SIMULACRO SABER 11"
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    
    ws['A3'] = f"Total Estudiantes: {len(datos)}"
    puntajes_g = [d['puntaje_global'] for d in datos if d['puntaje_global'] > 0]
    if puntajes_g:
        ws['A4'] = f"Promedio Global: {mean(puntajes_g):.1f}/500"
    else:
        ws['A4'] = "Promedio Global: N/A"
    
    # Distribucion por nivel
    row = 6
    ws[f'A{row}'] = "Distribucion por Nivel:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    for nivel in ['Superior', 'Alto', 'Medio', 'Bajo']:
        count = sum(1 for d in datos if d['nivel'] == nivel)
        pct = (count/len(datos)*100) if datos else 0
        ws[f'A{row}'] = f"  {nivel}: {count} ({pct:.1f}%)"
        row += 1
    
    # Hoja Todos los Estudiantes
    ws = wb.create_sheet("Todos los Estudiantes")
    headers = ["Institucion", "Correo", "Documento", "Nombres", "Apellidos",
               "Aciertos S1", "Aciertos S2", "Matematicas", "Lectura", 
               "Sociales", "Ciencias", "Ingles", "Puntaje Global", "Nivel"]
    aplicar_header(ws, 1, headers)
    
    for i, d in enumerate(sorted(datos, key=lambda x: x['puntaje_global'], reverse=True), 2):
        row_data = [
            d['institucion'], d['correo'], d['documento'], d['nombres'], d['apellidos'],
            f"{d['s1_aciertos']}/{d['s1_total']}", f"{d['s2_aciertos']}/{d['s2_total']}",
            f"{d['matematicas']}/100", f"{d['lectura']}/100", f"{d['sociales']}/100",
            f"{d['ciencias']}/100", f"{d['ingles']}/100", f"{d['puntaje_global']}/500", d['nivel']
        ]
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = thin_border
            if j == len(row_data):
                cell.fill = PatternFill(start_color=d['color'], end_color=d['color'], fill_type='solid')
    
    # Ajustar anchos
    for i, w in enumerate([20, 30, 15, 20, 20, 12, 12, 12, 12, 12, 12, 12, 15, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'F2'
    
    # Hojas por institucion
    por_inst = defaultdict(list)
    for d in datos:
        por_inst[d['institucion']].append(d)
    
    for inst, ests in sorted(por_inst.items()):
        safe_name = "".join([c for c in inst if c.isalnum() or c in ' _'])[:28] or "SinNombre"
        ws = wb.create_sheet(safe_name)
        
        ws.merge_cells('A1:N1')
        ws['A1'].value = f"INSTITUCION: {inst}"
        ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
        
        puntajes_inst = [e['puntaje_global'] for e in ests if e['puntaje_global'] > 0]
        prom_inst = mean(puntajes_inst) if puntajes_inst else 0
        ws['A3'] = f"Estudiantes: {len(ests)} | Promedio: {prom_inst:.0f}/500"
        
        aplicar_header(ws, 5, headers)
        
        aplicar_header(ws, 5, headers)
        
        # Calcular Indice Global (Top 80%) para header
        # Necesitamos reconstruir la estructura de estudiante completa o pasar datos compatibles
        # 'ests' es la lista de diccionarios planos (datos). Necesitamos reconstruir estructura 'puntajes'
        # para que funcione calcular_indice_global_colegio.
        # PERO 'generar_reporte_mejorado' recibe 'estudiantes_list' con estructura anidada.
        # 'd' en el bucle 'for d in datos' ya es plano.
        # Debemos buscar los estudiantes ORIGINALES asociados a esta institucion.
        
        # Filtrar de la lista original 'estudiantes_list' (que tiene la estructura anidada)
        ests_originales = [e for e in estudiantes_list if e['informacion_personal'].get('institucion', 'Sin Asignar') == inst]
        
        # Calcular Indice
        res_indice = calcular_indice_global_colegio(ests_originales)
        idx_val = res_indice['indice']
        cat_val = res_indice['clasificacion']
        n_top = res_indice['top_80_count']
        
        # Mostrar en el reporte (Fila 2 y 3, para evitar conflicto con Merge A1:N1)
        ws['I2'].value = f"INDICE GLOBAL (Top {n_top} est): {idx_val} / 13"
        ws['I2'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['I2'].fill = PatternFill(start_color=COLORS['header2'], end_color=COLORS['header2'], fill_type='solid')
        
        ws['I3'].value = f"CLASIFICACIÓN PROYECTADA: {cat_val}"
        ws['I3'].font = Font(bold=True, size=11, color='FFFFFF')
        ws['I3'].fill = PatternFill(start_color=COLORS['header2'], end_color=COLORS['header2'], fill_type='solid')
        
        for i, d in enumerate(sorted(ests, key=lambda x: x['puntaje_global'], reverse=True), 6):
            row_data = [
                d['institucion'], d['correo'], d['documento'], d['nombres'], d['apellidos'],
                f"{d['s1_aciertos']}/{d['s1_total']}", f"{d['s2_aciertos']}/{d['s2_total']}",
                f"{d['matematicas']}/100", f"{d['lectura']}/100", f"{d['sociales']}/100",
                f"{d['ciencias']}/100", f"{d['ingles']}/100", f"{d['puntaje_global']}/500", d['nivel']
            ]
            for j, val in enumerate(row_data, 1):
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = thin_border
                if j == len(row_data):
                    cell.fill = PatternFill(start_color=d['color'], end_color=d['color'], fill_type='solid')
    
    # ==========================================
    # HOJA: BASE DE DATOS
    # Diseño solicitado: Claro, explícito, datos necesarios
    # ==========================================
    ws = wb.create_sheet("Base de Datos")
    
    # Configuración de Headers
    headers_bd = [
        "#", "ID", "Nombres", "Apellidos", 
        "Global", "Matemáticas", "Lectura", "Sociales", "Ciencias", "Inglés", 
        "Nivel", "Institución", "S1", "S2"
    ]
    
    # 1. Aplicar Headers con diseño
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') # Azul oscuro profesional
    header_font = Font(bold=True, size=11, color='FFFFFF')
    
    for col_idx, header in enumerate(headers_bd, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.row_dimensions[1].height = 25 # Altura del header
    ws.freeze_panes = 'A2' # Congelar primera fila

    # 2. Llenar datos
    # Ordenar por puntaje global (mayor a menor)
    estudiantes_ordenados = sorted(datos, key=lambda x: x['puntaje_global'], reverse=True)
    
    for i, d in enumerate(estudiantes_ordenados, 2):
        # Preparar fila
        # S1 y S2 formato "Aciertos/Total"
        s1_str = f"{d['s1_aciertos']}/{d['s1_total']}"
        s2_str = f"{d['s2_aciertos']}/{d['s2_total']}"
        
        row_values = [
            i - 1, # # Index
            d['documento'],
            d['nombres'],
            d['apellidos'],
            d['puntaje_global'],
            d['matematicas'],
            d['lectura'],
            d['sociales'],
            d['ciencias'],
            d['ingles'],
            d['nivel'],
            d['institucion'],
            s1_str,
            s2_str
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Alineación especial para nombres (Izquierda)
            if col_idx in [3, 4, 12]: # Nombres, Apellidos, Institución
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Formato Condicional para Global (Col 5)
            if col_idx == 5:
                cell.font = Font(bold=True)
                if val >= 400: cell.fill = PatternFill(start_color=COLORS['superior'], end_color=COLORS['superior'], fill_type='solid')
                elif val >= 325: cell.fill = PatternFill(start_color=COLORS['alto'], end_color=COLORS['alto'], fill_type='solid')
                elif val >= 250: cell.fill = PatternFill(start_color=COLORS['medio'], end_color=COLORS['medio'], fill_type='solid')
                else: cell.fill = PatternFill(start_color=COLORS['bajo'], end_color=COLORS['bajo'], fill_type='solid')

            # Formato Condicional para Áreas (Col 6-10)
            if 6 <= col_idx <= 10:
                 if val >= 80: cell.fill = PatternFill(start_color=COLORS['superior'], end_color=COLORS['superior'], fill_type='solid')
                 elif val >= 65: cell.fill = PatternFill(start_color=COLORS['alto'], end_color=COLORS['alto'], fill_type='solid')
                 elif val >= 50: cell.fill = PatternFill(start_color=COLORS['medio'], end_color=COLORS['medio'], fill_type='solid')
                 else: cell.fill = PatternFill(start_color=COLORS['bajo'], end_color=COLORS['bajo'], fill_type='solid')

        # Alternar color de fila para legibilidad (Zebra striping)
        if i % 2 == 0: # Filas pares
            for col_idx in range(1, len(row_values) + 1):
                cell = ws.cell(row=i, column=col_idx)
                # Aplicar solo si no tiene color de fondo ya
                if cell.fill and cell.fill.fgColor.rgb != '00000000': 
                    continue # Saltarse celdas con color existente
                
                cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')

    # 3. Ajustar anchos de columna automáticamente
    col_widths = {
        1: 5,   # #
        2: 15,  # ID
        3: 25,  # Nombres
        4: 25,  # Apellidos
        5: 10,  # Global
        6: 12,  # Mat
        7: 12,  # Lec
        8: 12,  # Soc
        9: 12,  # Cie
        10: 12, # Ing
        11: 15, # Nivel
        12: 30, # Institución
        13: 12, # S1
        14: 12  # S2
    }
    
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Filtros automáticos
    ws.auto_filter.ref = ws.dimensions
    
    output = SALIDA_DIR / "Reporte_Completo.xlsx"
    wb.save(output)
    return output

# ============================================================================
# FUNCION PRINCIPAL
# ============================================================================
def main():
    print("=" * 60)
    print("SISTEMA DE PROCESAMIENTO - SIMULACRO SABER 11")
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    # Detectar archivos
    archivos = detectar_archivos()
    
    missing = [k for k, v in archivos.items() if v is None]
    if missing:
        print(f"\n[!] Archivos faltantes: {missing}")
        print("    Coloca los archivos en entrada/respuestas/ y entrada/claves/")
        return
    
    # Cargar configuración de invalidaciones
    invalidaciones = cargar_invalidaciones()
    if invalidaciones:
        print(f"\n      [!] Se cargaron {len(invalidaciones)} pregunta(s) invalidada(s):")
        for inv in invalidaciones:
            print(f"          - {inv.get('sesion', '?')}: Pregunta {inv.get('numero_pregunta', '?')} ({inv.get('motivo', 'Sin motivo')})")
    
    # Procesar sesiones
    print("\n[2/4] Procesando sesiones...")
    resultados_s1 = procesar_sesion(archivos['respuestas_s1'], archivos['claves_s1'], 'S1', invalidaciones)
    print(f"      S1: {len(resultados_s1)} estudiantes")
    
    resultados_s2 = procesar_sesion(archivos['respuestas_s2'], archivos['claves_s2'], 'S2', invalidaciones)
    print(f"      S2: {len(resultados_s2)} estudiantes")
    
    # Combinar resultados usando múltiples identificadores
    print("\n[3/4] Combinando datos...")
    estudiantes = {}
    
    # Índices para buscar por múltiples claves
    indice_correo = {}
    indice_nombre = {}
    indice_id = {}
    
    # Primero, agregar todos los de S1 y crear índices
    for id_est, datos in resultados_s1.items():
        estudiantes[id_est] = datos
        
        info = datos['informacion_personal']
        correo = normalizar_correo(info.get('correo_electronico', ''))
        nombre_key = normalizar_nombre(info.get('nombres', ''), info.get('apellidos', ''))
        
        if correo and '@' in correo:
            indice_correo[correo] = id_est
        if nombre_key and len(nombre_key) > 3:
            indice_nombre[nombre_key] = id_est
        indice_id[id_est] = id_est
    
    # Combinar S2 usando múltiples estrategias de matching
    matches_encontrados = 0
    for id_est, datos in resultados_s2.items():
        info = datos['informacion_personal']
        correo = normalizar_correo(info.get('correo_electronico', ''))
        nombre_key = normalizar_nombre(info.get('nombres', ''), info.get('apellidos', ''))
        
        # Buscar match en múltiples índices
        match_id = None
        match_tipo = None
        
        # 1. Buscar por correo (más confiable)
        if correo and correo in indice_correo:
            match_id = indice_correo[correo]
            match_tipo = 'correo'
        # 2. Buscar por ID exacto
        elif id_est in indice_id:
            match_id = indice_id[id_est]
            match_tipo = 'id'
        # 3. Buscar por nombre+apellido
        elif nombre_key and nombre_key in indice_nombre:
            match_id = indice_nombre[nombre_key]
            match_tipo = 'nombre'
        
        if match_id and match_id in estudiantes:
            # Encontrado - combinar datos
            matches_encontrados += 1
            estudiantes[match_id]['s2_aciertos'] = datos['s2_aciertos']
            estudiantes[match_id]['s2_total'] = datos['s2_total']
            for mat, punt in datos['puntajes'].items():
                if mat in estudiantes[match_id]['puntajes']:
                    estudiantes[match_id]['puntajes'][mat]['correctas'] += punt['correctas']
                    estudiantes[match_id]['puntajes'][mat]['total_preguntas'] += punt['total_preguntas']
                else:
                    estudiantes[match_id]['puntajes'][mat] = punt
            if 'S2' not in estudiantes[match_id]['sesiones']:
                estudiantes[match_id]['sesiones'].append('S2')
            # Actualizar respuestas detalladas
            for mat, resp in datos.get('respuestas_detalladas', {}).items():
                if 'respuestas_detalladas' not in estudiantes[match_id]:
                    estudiantes[match_id]['respuestas_detalladas'] = {}
                if mat not in estudiantes[match_id]['respuestas_detalladas']:
                    estudiantes[match_id]['respuestas_detalladas'][mat] = resp
                else:
                    estudiantes[match_id]['respuestas_detalladas'][mat].extend(resp)
        else:
            # No encontrado - agregar como nuevo
            estudiantes[id_est] = datos
    
    print(f"      Matches encontrados: {matches_encontrados}")
    
    # Calcular puntajes finales
    calcular_puntajes_finales(estudiantes)
    
    ambas = sum(1 for e in estudiantes.values() if 'S1' in e['sesiones'] and 'S2' in e['sesiones'])
    print(f"      Total: {len(estudiantes)} estudiantes ({ambas} con ambas sesiones)")
    
    # Guardar JSON
    lista_est = list(estudiantes.values())
    resultado = {
        'estudiantes': lista_est,
        'metadata': {
            'fecha_generacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_estudiantes': len(lista_est)
        }
    }
    
    json_path = SALIDA_DIR / "resultados_finales.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    
    # Generar reportes
    print("\n[4/4] Generando reportes...")
    reporte = generar_reporte_mejorado(lista_est)
    
    print("\n" + "=" * 60)
    print("[OK] PROCESAMIENTO COMPLETADO")
    print("=" * 60)
    print(f"\nArchivos generados en: {SALIDA_DIR}")
    print(f"  - resultados_finales.json")
    print(f"  - {reporte.name}")
    print("[OK] PROCESAMIENTO COMPLETADO")
    return resultado

if __name__ == '__main__':
    main()
